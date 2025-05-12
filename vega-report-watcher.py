import sqlite3 # библиотека для чтения базы данных sqlite (server.db)
from datetime import datetime # библиотека для преобразования unix_time в нормальное представление
from openpyxl import Workbook # библиотека для формирования xlsx документа
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from pathlib import Path
from time import sleep
import os
import sys
import json

version = '1.0.0.2 (от 12.05.2025)'
version_name = 'верба'

# лист со списком поддерживаемых устройств
support_device_list = ['СИ-12', 'Smart Badge', 'ТС-12', 'Smart-WB0101', 'Smart-HS0101', 'Smart-MS0101', 'ТД-11v2', 'Smart-MC0101', 'Smart-SS0102', 'Smart-UM0101']


# настройки работы приложения
settings_app = [0, '', '', 0] # отображать ошибки (0/1), отображать только отчёты указанного устройства (''/'Устройство'), отображать только для определенного deveui (''/'DEVEUI), выводить все или только новые отчёты (0 - все , 1 - только новые)


print(f'[Vega Report Watcher]\nДанная программа создана для предоставления отчетов устройств в понятном для любого человека виде.\nВ данный момент поддерживаются следующие виды устройств: {", ".join(support_device_list)}.\nВерсия программы: {version} (кодовое наименование: {version_name}).')


# базовое окно Tkinter
root = Tk()
root.withdraw()

# функция преобразования данных в нормальное представление (little) (10-ое) [hex -> dec]
def human_watch(hex_data, sign=False):
    value = int.from_bytes(hex_data, byteorder='little', signed=sign) # если больше чем 1 байт, то от младшего байта к старшему (СБ <- МБ или df <- a0 <- 01 <- 2f (влево движение))
    return value
# функция преобразования данных в нормальное представление (big) (10-ое) [hex -> dec]
def human_watch_big(hex_data, sign=False):
    value = int.from_bytes(hex_data, byteorder='big', signed=sign) # если больше чем 1 байт, то от старшего байта к младшему (СБ -> МБ или df -> a0 -> 01 -> 2f (вправо движение))
    return value

# функция для преобразования географических координат в десятичный вид
def dms_to_decimal(degrees: int, minutes: float, direction: str = '') -> float:
    decimal = degrees + minutes / 60
    if direction in ['S', 'W']:
        decimal *= -1
    return decimal

# функция поиска соответствующей записи в листе
def find_index(mass, key):
    for i, item in enumerate(mass):
        if item[0] == key:
            return i
    return -1  # или None, если не найдено


start_time = datetime.now() # время запуска проверки
count_retry_select_file_path = 0
while True:
    print('Выберите файл server.db через диалог выбора файла...')
    file_path = askopenfilename(
        title="Выберите файл server.db",
        filetypes=[("SQLite Database", "server.db")],
        initialfile="server.db"
    )

    if not file_path:
        print(f"Файл не выбран. Оставшееся число попыток: {5 - (count_retry_select_file_path + 1)}. Повтор попытки...")
        if (count_retry_select_file_path == 4):
            print('Превышено количество попыток выбора файла.\nПрограмма будет закрыта через 2 секунды.')
            sleep(2)
            exit(1)
        count_retry_select_file_path += 1
        continue

    path = Path(file_path)

    if path.exists() and path.is_file() and path.suffix == ".db" and path.name == "server.db":
        print(f'Файл успешно выбран: {path}')
        break
    else:
        print(f'Неверный файл. Убедитесь, что выбран файл server.db. Повтор попытки...')

# подключение к файлу базы данных
connection = sqlite3.connect(file_path)
# создание проводника по базе данных
cursor = connection.cursor()

devices_DB = cursor.execute(f"SELECT devname, deveui FROM devices").fetchall()
devices_list = []
for i in devices_DB:
    if (i[1] == ''):
        continue
    try:
        classroom = f'{i[0]}'.split(' ')[1]
        if (classroom == 'Badge'):
            devices_list.append([f'{i[1]}', f'Smart Badge', f'ауд. неизвестно'])
        else:
            devices_list.append([f'{i[1]}', f'{i[0]}'.split(' ')[0], f'ауд. {classroom}'])
    except:
        devices_list.append([f'{i[1]}', f'{i[0]}'.split(' ')[0], f'ауд. неизвестно'])

bs_DB = cursor.execute(f"SELECT mac, comment, hostaddress FROM bs").fetchall()
bs_list = []
for i in bs_DB:
    name = json.loads(i[1])['name']
    bs_list.append([f'{i[0]}', f'{name}'.split('_')[0], i[2]])

print('[Генерация динамических массивов устройств и базовых станций завершена]')

print('\n[Настройка параметров визуальной обработки отчётов]')

variable_view_error = input('Хотите видеть ошибки при обработке отчётов? (стандартно - нет) (1 - да, 0 или любой другой символ (также просто оставить пустым при нажатии enter) - нет): ')
if (variable_view_error == '1'):
    settings_app[0] = 1
else:
    settings_app[0] = 0
while True:
    variable_view_fix_deveui = (input('\nХотите посмотреть отчёты какого-то конкретного устройства? (введите его DEVEUI в любом виде) [введите ничего или 0 - не указывать, введите 1 - показать список поддерживаемых устройств]: ')).upper()
    if (variable_view_fix_deveui == '0' or variable_view_fix_deveui == ''):
        break
    if (variable_view_fix_deveui == '1'):
        print('DEVEUI - Тип устройства')
        for i in devices_list:
            print(f'{i[0]} - {i[1]} ({i[2]})')
        continue
    response = find_index(devices_list, variable_view_fix_deveui)
    if (response == None or response == -1):
        print(f'Введённый DEVEUI не найден в списке поддерживаемых устройств, добавьте его самостоятельно или попробуйте ввести повторно.')
        continue
    else:
        settings_app[2] = variable_view_fix_deveui
        break
variable_view_reports_for_time = input('\nХотите видеть только новые отчёты? (стандартно - нет) (1 - да, 0 или любой другой символ (также просто оставить пустым при нажатии enter) - нет): ')
if (variable_view_reports_for_time == '1'):
    settings_app[3] = 1
else:
    settings_app[3] = 0
# массив с проверенными отчётами (лежит время отчётов)
processed_time_list = []

"""
file_path = 'server.db'
while True:
    path = Path(input('Введите путь до файла базы данных (server.db) (если программа находится в одной папке с файлом БД, можно пропустить ввод): '))
    if (str(path) == '.'):
        path = Path("server.db")
        if path.exists() and path.is_file() and path.suffix == ".db" and path.name == "server.db":
            break
        else:
            print('Файл server.db не найден в одной директории с программой.')
            continue
    else:
        path = Path(f"{path}\server.db")
        if path.exists() and path.is_file() and path.suffix == ".db" and path.name == "server.db":
            file_path = path
            break
        else:
            print(f'Файл server.db не найден по указанному пути ({path}).')
            continue
            
"""

print('[Настройки применены]')
sleep(1)

# базовое формирование документа xlsx
wb = Workbook()
ws = wb[f'Sheet']
ws.title = "Основная информация"
ws.cell(row=2, column=2, value=f'В данном документе для каждого устройства создан отдельный лист формата DEVEUI (DEVICE_TYPE)')
ws.cell(row=2, column=3, value=f'В этих листах хранятся отформатированные отчёты из таблицы (rawdata) из базы данных сервера Веги')
ws.cell(row=2, column=4, value=f'Форматы столбцов разнятся, но у всех первый столбец обозначает время принятия отчёта на сервере и все отчёты идут сверху вниз, от старых к новым')

for i in devices_list:
    wb.create_sheet(f'{i[0]} ({i[1]})')
    ws = wb[f'{i[0]} ({i[1]})']
    if (i[1] == 'Smart-MS0101'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Тип пакета')
        ws.cell(row=2, column=4, value=f'Заряд батареи, %')
        ws.cell(row=2, column=5, value=f'Основные параметры')
        ws.cell(row=2, column=6, value=f'Температура, *C')
        ws.cell(row=2, column=7, value=f'Причина отправки пакета')
        ws.cell(row=2, column=8, value=f'Время')
    elif (i[1] == 'ТС-12'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Биты присутствия')
        ws.cell(row=2, column=4, value=f'Причина отправки пакета')
        ws.cell(row=2, column=5, value=f'Координаты')
        ws.cell(row=2, column=6, value=f'Количество принятых пакетов')
        ws.cell(row=2, column=7, value=f'Количество отправленных пакетов')
        ws.cell(row=2, column=8, value=f'Заряд батареи, мВ')
        ws.cell(row=2, column=9, value=f'RSSI, dBm')
        ws.cell(row=2, column=10, value=f'SNR, dB')
    elif (i[1] == 'Smart-WB0101'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Режим работы')
        ws.cell(row=2, column=4, value=f'Заряд батареи, %')
        ws.cell(row=2, column=5, value=f'Время')
        ws.cell(row=2, column=6, value=f'Температура')
    elif (i[1] == 'СИ-12'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Тип пакета')
        ws.cell(row=2, column=4, value=f'Заряд батареи, %')
        ws.cell(row=2, column=5, value=f'Основные параметры')
        ws.cell(row=2, column=6, value=f'Время')
        ws.cell(row=2, column=7, value=f'Температура, *C')
        ws.cell(row=2, column=8, value=f'Показание на входе 1')
        ws.cell(row=2, column=9, value=f'Показание на входе 2')
        ws.cell(row=2, column=10, value=f'Показание на входе 3')
        ws.cell(row=2, column=11, value=f'Показание на входе 4')
    elif (i[1] == 'Smart-HS0101'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Тип пакета')
        ws.cell(row=2, column=4, value=f'Заряд батареи, %')
        ws.cell(row=2, column=5, value=f'Время')
        ws.cell(row=2, column=6, value=f'Температура, *C')
        ws.cell(row=2, column=7, value=f'Влажность, %')
        ws.cell(row=2, column=8, value=f'Состояние датчика открытия 1')
        ws.cell(row=2, column=9, value=f'Состояние датчика открытия 2')
        ws.cell(row=2, column=10, value=f'Угол отклонения от вертикали, *')
        ws.cell(row=2, column=11, value=f'Нижний порог влажности, %')
        ws.cell(row=2, column=12, value=f'Верхний порог влажности, %')
        ws.cell(row=2, column=13, value=f'Нижний порог температуры, *C')
        ws.cell(row=2, column=14, value=f'Верхний порог температуры, *C')
    elif (i[1] == 'Smart Badge'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Тип пакета')
        ws.cell(row=2, column=4, value=f'Причина формирования пакета')
        ws.cell(row=2, column=5, value=f'Заряд батареи, %')
        ws.cell(row=2, column=6, value=f'Время')
        ws.cell(row=2, column=7, value=f'Температура, *C')
        ws.cell(row=2, column=8, value=f'Угол отклонения от вертикали, *')
        # для типа 1
        ws.cell(row=2, column=9, value=f'Широта (только тип 1)')
        ws.cell(row=2, column=10, value=f'Долгота (только тип 1)')
        ws.cell(row=2, column=11, value=f'Курс, * (только тип 1)')
        ws.cell(row=2, column=12, value=f'Скорость, км/ч (только тип 1)')
        ws.cell(row=2, column=13, value=f'Высота над ср. уровнем моря, м. (только тип 1)')
        ws.cell(row=2, column=14, value=f'Количество видимых спутников (только тип 1)')
        ws.cell(row=2, column=15, value=f'Количество спутников решения (только тип 1)')
        # для типа 2
        ws.cell(row=2, column=16, value=f'Тип BLE-маяка (только тип 2)')
        ws.cell(row=2, column=17, value=f'Наименование BLE-маяка (только тип 2)')
        ws.cell(row=2, column=18, value=f'RSSI эталонное, dBm (только тип 2)')
        ws.cell(row=2, column=19, value=f'TX_POWER (только тип 2)')
        # для типа 5
        ws.cell(row=2, column=20, value=f'MAC для BLE-меток №1,2,3 (только тип 5)')
        ws.cell(row=2, column=21, value=f'Заряд батареи для BLE-меток №1,2,3, % (только тип 5)')
        ws.cell(row=2, column=22, value=f'Температура для BLE-меток №1,2,3, *C (только тип 5)')
        ws.cell(row=2, column=23, value=f'Влажность для BLE-меток №1,2,3, % (только тип 5)')
        ws.cell(row=2, column=24, value=f'RSSI эталонное для BLE-меток №1,2,3, dBm (только тип 5)')
        ws.cell(row=2, column=25, value=f'TX_POWER для BLE-меток №1,2,3 (только тип 5)')
        # общая 
        ws.cell(row=2, column=26, value=f'Состояние меток СИЗ')
    elif (i[1] == 'ТД-11v2'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Тип пакета')
        ws.cell(row=2, column=4, value=f'Заряд батареи, %')
        ws.cell(row=2, column=5, value=f'Состояние лимитов')
        ws.cell(row=2, column=6, value=f'Время')
        ws.cell(row=2, column=7, value=f'Температура устройства, *C')
        ws.cell(row=2, column=8, value=f'Температура датчика NTC, *C')
        ws.cell(row=2, column=9, value=f'Нижний предел температуры датчка, *C')
        ws.cell(row=2, column=10, value=f'Верхний предел температуры датчка, *C')
        ws.cell(row=2, column=11, value=f'Состояние дискретного входа')
        ws.cell(row=2, column=12, value=f'Показатели на дискретном входе')
    elif (i[1] == 'Smart-MC0101'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value=f'Тип пакета')
        ws.cell(row=2, column=4, value=f'Заряд батареи, %')
        ws.cell(row=2, column=5, value=f'Температура, *C')
        ws.cell(row=2, column=6, value=f'Причина отправки пакета')
        ws.cell(row=2, column=7, value=f'Состояние входов')
        ws.cell(row=2, column=8, value=f'Время снятия показаний/Время формирования пакета')
    elif(i[1] == 'Smart-SS0102'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value='Тип пакета')
        ws.cell(row=2, column=4, value='Время формирования пакета')
        ws.cell(row=2, column=5, value='Текущий статус')
        ws.cell(row=2, column=6, value='Напряжение с приемника, мВ')
        ws.cell(row=2, column=7, value='Ток передатчика, мА')
        ws.cell(row=2, column=8, value='Температура на термисторе, *C')
        ws.cell(row=2, column=9, value='Флаг - используется батарея 1')
        ws.cell(row=2, column=10, value='Флаг - используется батарея 2')
        ws.cell(row=2, column=11, value='Флаг присутствия батареи 1')
        ws.cell(row=2, column=12, value='Флаг присутствия батареи 2')
        ws.cell(row=2, column=13, value='Заряд батареи 1, %')
        ws.cell(row=2, column=14, value='Заряд батареи 2, %')
    elif(i[1] == 'Smart-UM0101'):
        ws.cell(row=2, column=1, value=f'Время отчёта')
        ws.cell(row=2, column=2, value=f'Принято с БС')
        ws.cell(row=2, column=3, value='Тип пакета')
        ws.cell(row=2, column=4, value='Заряд батареи, %')
        ws.cell(row=2, column=5, value='Время снятия показаний')
        ws.cell(row=2, column=6, value='Состояние питания')
        ws.cell(row=2, column=7, value='Температура, *C')
        ws.cell(row=2, column=8, value='Влажность, %')
        ws.cell(row=2, column=9, value='Уровень освещенности')
        ws.cell(row=2, column=10, value='Уровень шума')
        ws.cell(row=2, column=11, value='Уровень CO2, ppm')
        ws.cell(row=2, column=12, value='Угол отклонения от вертикали')
        ws.cell(row=2, column=13, value='Нижний порог температуры, *C')
        ws.cell(row=2, column=14, value='Верхний порог температуры, *C')
        ws.cell(row=2, column=15, value='Нижний порог влажности, %')
        ws.cell(row=2, column=16, value='Верхний порог влажности, %')
        ws.cell(row=2, column=17, value='Нижний порог уровня освещенности')
        ws.cell(row=2, column=18, value='Верхний порог уровня освещенности')
        ws.cell(row=2, column=19, value='Нижний порог уровня шума')
        ws.cell(row=2, column=20, value='Верхний порог уровня шума')
        ws.cell(row=2, column=21, value='Нижний порог уровня CO2, ppm')
        ws.cell(row=2, column=22, value='Верхний порог уровня CO2, ppm')
    ws.cell(row=1, column=1, value='DEVEUI: ')
    ws.cell(row=1, column=3, value='Название устройства: ')
    ws.cell(row=1, column=5, value='Аудитория: ')


try:
    print('[Процесс обработки]: Обработка уже полученных ранее отчётов.')
    count = 0
    while True:

        # получение данных из таблицы rawdata с колонок data, port, deveui, time
        data_DB = cursor.execute(f"SELECT data, port, deveui, time, macbs FROM rawdata").fetchall()

        for i in range(len(data_DB) - 1):
            if (i == ''): # на всякий пропуск пустой записи (если будет)
                continue
            data_raw = data_DB[i][0]
            port = data_DB[i][1]
            unix_time = data_DB[i][3]
            if (unix_time in processed_time_list):
                continue
            else:
                processed_time_list.append(unix_time)
            timestamp_bd= datetime.fromtimestamp(unix_time * 1e-3)
            deveui = data_DB[i][2]
            macbs_buff = str(data_DB[i][4]).split('+')
            macbs = ''
            for i in macbs_buff:
                if i != macbs_buff[len(macbs_buff) -1]:
                    macbs += f'{i} ({bs_list[find_index(bs_list, i)][1]}) [{bs_list[find_index(bs_list, i)][2]}] + '
                else:
                    macbs += f'{i} ({bs_list[find_index(bs_list, i)][1]}) [{bs_list[find_index(bs_list, i)][2]}].'


            if (find_index(devices_list, deveui) == -1):
                if(settings_app[3] == 0):
                    print(f'[Ошибка в отчёте {timestamp_bd}]: К сожалению, данное устройство: "{deveui}" пока не поддерживается\n= = =')
                continue
            type_device = devices_list[find_index(devices_list, deveui)][1]
            classroom = devices_list[find_index(devices_list, deveui)][2]
            if (settings_app[1] != ''):
                if (type_device != settings_app[1]):
                    continue
            """
            if (settings_app[2] != ''):
                if (deveui != settings_app[2]):
                    continue
            """
                

            # Smart Badge Вега
            if (type_device == 'Smart Badge' and port == 2): # формат для данных, приходящих на 2 порт LoRaWAN!!!
                type_packet = human_watch(data_raw[:1])
                if (type_packet == 1):
                    type_packet_decode = 'определение координат по GPS/ГЛОНАСС'
                elif (type_packet == 2):
                    type_packet_decode = 'определение ID ближайшей BLE-метки'
                elif (type_packet == 5):
                    type_packet_decode = 'определение трёх ближайших BLE-меток Вега'
                else:
                    type_packet_decode = f'отличный от 1,2,5 типа: {type_packet}'
                reason_form_packet = human_watch(data_raw[1:2])
                if (reason_form_packet == 0):
                    reason_format_packet_decode = 'по времени'
                elif (reason_form_packet == 1):
                    reason_format_packet_decode = 'по началу движения'
                elif (reason_form_packet == 2):
                    reason_format_packet_decode = 'по прекращению движения'
                elif (reason_form_packet == 3):
                    reason_format_packet_decode = 'по датчику отрыва'
                elif (reason_form_packet == 4):
                    reason_format_packet_decode = 'по обнаружению падения (удар)'
                elif (reason_form_packet == 5):
                    reason_format_packet_decode = 'по активации тревоги (поиск)'
                elif (reason_form_packet == 6):
                    reason_format_packet_decode = 'по активации тревоги (потеря меток СИЗ)'
                battery = human_watch(data_raw[2:3])
                unix_time = human_watch(data_raw[3:7])
                timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                temperature = human_watch(data_raw[7:8], True)
                state = f'{human_watch(data_raw[8:9]):08b}'
                state_string = ''
                if (state[0] == '0'):
                    state_string += 'Признак движения: нет\n'
                else:
                    state_string += 'Признак движения: да\n'
                if (state[1] == '0'):
                    state_string += 'Признак фиксации падения: нет\n'
                else:
                    state_string += 'Признак фиксации падения: да\n'
                if (state[2] == '0'):
                    state_string += 'Признак валидности координат (для GPS): не валидны\n'
                else:
                    state_string += 'Признак валидности координат (для GPS): валидны\n'
                # тип активной команды
                if (state[3] == '0'):
                    if (state[4] == '0'):
                        if (state[5] == '0'):
                            state_string += 'Тип активной команды: нет активной команды.'
                        else:
                            state_string += 'Тип активной команды: резерв.'
                    else:
                        if (state[5] == '0'):
                            state_string += 'Тип активной команды: предупреждение.'
                        else:
                            state_string += 'Тип активной команды: резерв.'
                else:
                    if (state[4] == '0'):
                        if (state[5] == '0'):
                            state_string += 'Тип активной команды: вызов.'
                        else:
                            state_string += 'Тип активной команды: резерв.'
                    else:
                        if (state[5] == '0'):
                            state_string += 'Тип активной команды: поиск.'
                        else:
                            state_string += 'Тип активной команды: отмена.'
                ugol_otkl_vert = human_watch(data_raw[9:11])
                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=f'{type_packet_decode} ({type_packet})')
                ws.cell(row=last_row+1, column=4, value=reason_format_packet_decode)
                ws.cell(row=last_row+1, column=5, value=battery)
                ws.cell(row=last_row+1, column=6, value=timestamp)
                ws.cell(row=last_row+1, column=7, value=temperature)
                ws.cell(row=last_row+1, column=8, value=ugol_otkl_vert)

                if (type_packet == 1):
                    latitude = int(human_watch(data_raw[11:15], True)) / 1000000
                    longtitude = int(human_watch(data_raw[15:19], True)) / 1000000
                    course = human_watch(data_raw[19:21])
                    speed = human_watch(data_raw[21:23])
                    height = human_watch(data_raw[23:25], True)
                    count_sputnik = human_watch(data_raw[25:26])
                    count_decision_sputnik = human_watch(data_raw[26:27])
                    state_tag_SIZ = f'{human_watch(data_raw[27:28]):08b}'
                    if (state_tag_SIZ[:3] == '000'):
                        state_tag_SIZ_decode = 'нет потерянных меток'
                    elif (state_tag_SIZ[:3] == '111'):
                        state_tag_SIZ_decode = 'все метки потеряны'
                    else:
                        state_tag_SIZ_decode = f'потеряны метки, у которых значение 1: {state_tag_SIZ[:3]} (нумерация с метки №1)'
                    ws.cell(row=last_row+1, column=9, value=latitude)
                    ws.cell(row=last_row+1, column=10, value=longtitude)
                    ws.cell(row=last_row+1, column=11, value=course)
                    ws.cell(row=last_row+1, column=12, value=speed)
                    ws.cell(row=last_row+1, column=13, value=height)
                    ws.cell(row=last_row+1, column=14, value=count_sputnik)
                    ws.cell(row=last_row+1, column=15, value=count_decision_sputnik)
                    ws.cell(row=last_row+1, column=26, value=state_tag_SIZ_decode)

                if (type_packet == 2):
                    type_ble_beacon = human_watch(data_raw[11:12], True)
                    if (type_ble_beacon == 0):
                        ble_beacon = 'маяк не обнаружен'
                    elif (type_ble_beacon == 1):
                        ble_beacon = 'iBeacon\n'
                        ble_beacon += f'UUID ({human_watch(data_raw[12:28])}) + Major ID ({human_watch(data_raw[28:30])})+ Minor ID ({human_watch(data_raw[30:32])})'
                    elif (type_ble_beacon == 2):
                        ble_beacon = 'Eddystone\n'
                        ble_beacon += f'Namespace ID ({human_watch(data_raw[12:22])}) + Instance ID с зап. нулями ({human_watch(data_raw[22:28])})'
                    elif (type_ble_beacon == 3):
                        ble_beacon = 'ALTBeacon\n'
                        ble_beacon += f'Beacon ID ({human_watch(data_raw[12:32])})'
                    elif (type_ble_beacon == 4):
                        ble_beacon = 'Вега\n'
                        ble_beacon += f'MAC ({human_watch(data_raw[12:18])}) + Заряд батареи ({human_watch(data_raw[18:19])} %) + Температура ({human_watch(data_raw[19:21])}*) + Влажность ({human_watch(data_raw[21:22])})'
                    rssi_etalon = human_watch(data_raw[32:33], True)
                    tx_power = human_watch(data_raw[33:34], True)
                    state_tag_SIZ = f'{human_watch(data_raw[34:35]):08b}'
                    if (state_tag_SIZ[:3] == '000'):
                        state_tag_SIZ_decode = 'нет потерянных меток'
                    elif (state_tag_SIZ[:3] == '111'):
                        state_tag_SIZ_decode = 'все метки потеряны'
                    else:
                        state_tag_SIZ_decode = f'потеряны метки, у которых значение 1: {state_tag_SIZ[:3]} (нумерация с метки №1)'
                    ws.cell(row=last_row+1, column=16, value=type_ble_beacon)
                    ws.cell(row=last_row+1, column=17, value=ble_beacon)
                    ws.cell(row=last_row+1, column=18, value=rssi_etalon)
                    ws.cell(row=last_row+1, column=19, value=tx_power)
                    ws.cell(row=last_row+1, column=26, value=state_tag_SIZ_decode)

                if (type_packet == 5):
                    # tag one
                    mac_tag_one = human_watch(data_raw[11:17])
                    battery_tag_one = human_watch(data_raw[17:18])
                    temperature_tag_one = human_watch(data_raw[18:19], True)
                    humidity_tag_one = human_watch(data_raw[19:20])
                    rssi_etalon_tag_one = human_watch(data_raw[20:21], True)
                    tx_power_tag_one = human_watch(data_raw[21:22], True)
                    # tag two
                    mac_tag_two = human_watch(data_raw[22:28])
                    battery_tag_two = human_watch(data_raw[28:29])
                    temperature_tag_two = human_watch(data_raw[29:30], True)
                    humidity_tag_two = human_watch(data_raw[30:31])
                    rssi_etalon_tag_two = human_watch(data_raw[31:32], True)
                    tx_power_tag_two = human_watch(data_raw[32:33], True)
                    # tag three
                    mac_tag_three = human_watch(data_raw[33:39])
                    battery_tag_three = human_watch(data_raw[39:40])
                    temperature_tag_three = human_watch(data_raw[40:41], True)
                    humidity_tag_three = human_watch(data_raw[41:42])
                    rssi_etalon_tag_three = human_watch(data_raw[42:43], True)
                    tx_power_tag_three = human_watch(data_raw[43:44], True)
                    state_tag_SIZ = f'{human_watch(data_raw[44:45]):08b}'
                    if (state_tag_SIZ[:3] == '000'):
                        state_tag_SIZ_decode = 'нет потерянных меток'
                    elif (state_tag_SIZ[:3] == '111'):
                        state_tag_SIZ_decode = 'все метки потеряны'
                    else:
                        state_tag_SIZ_decode = f'потеряны метки, у которых значение 1: {state_tag_SIZ[:3]} (нумерация с метки №1)'
                    ws.cell(row=last_row+1, column=20, value=f'{mac_tag_one}, {mac_tag_two}, {mac_tag_three}')
                    ws.cell(row=last_row+1, column=21, value=f'{battery_tag_one}, {battery_tag_two}, {battery_tag_three}')
                    ws.cell(row=last_row+1, column=22, value=f'{temperature_tag_one}, {temperature_tag_two}, {temperature_tag_three}')
                    ws.cell(row=last_row+1, column=23, value=f'{humidity_tag_one}, {humidity_tag_two}, {humidity_tag_three}')
                    ws.cell(row=last_row+1, column=24, value=f'{rssi_etalon_tag_one}, {rssi_etalon_tag_two}, {rssi_etalon_tag_three}')
                    ws.cell(row=last_row+1, column=25, value=f'{tx_power_tag_one}, {tx_power_tag_two}, {tx_power_tag_three}')
                    ws.cell(row=last_row+1, column=26, value=state_tag_SIZ_decode)
            
            # Кнопка спасения Smart-WB0101 Вега
            elif (type_device == 'Smart-WB0101' and port == 2): # формат для данных, приходящих на 2 порт LoRaWAN!!!
                operating_mode = human_watch(data_raw[:1])
                if (operating_mode == 1):
                    operating_mode_decode = 'ожидание'
                elif (operating_mode == 2):
                    operating_mode_decode = 'передача тревоги'
                elif (operating_mode == 3):
                    operating_mode_decode = 'тревога получена сервером'
                elif (operating_mode == 4):
                    operating_mode_decode = 'отмена тревоги'
                elif (operating_mode == 5):
                    operating_mode_decode = 'тревога принята оператором'
                else:
                    operating_mode_decode = 'нажатие, включен режим простой'
                battery = human_watch(data_raw[1:2])
                unix_time = human_watch(data_raw[2:6])
                timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                temperature = human_watch(data_raw[6:7])
                
                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=operating_mode_decode)
                ws.cell(row=last_row+1, column=4, value=battery)
                ws.cell(row=last_row+1, column=5, value=timestamp)
                ws.cell(row=last_row+1, column=6, value=temperature)
                
            # Тестер сети ТС-12 Вега
            elif (type_device == 'ТС-12' and port == 4): # формат для данных, приходящих на 4 порт LoRaWAN!!!
                bits_presence = f'{human_watch(data_raw[:1]):08b}'
                # температура в дополнительном коде (каком?)
                if (bits_presence[2] == '0'):
                    reason_send_packet = 'в автоматическом режиме'
                else:
                    reason_send_packet = 'по нажатию на кнопку'
                if (bits_presence[3] == '0'):
                    coordinates = 'навигационных данных нет.'
                    next_start_position = 1
                else:
                    coordinates = ''
                    latitude_degree = f'{human_watch_big(data_raw[1:2]):08b}'
                    buff_latitude_degree = f'{int(latitude_degree[:4], 2)}{int(latitude_degree[4:], 2)}'
                    coordinates += f'{int(latitude_degree[:4], 2)}{int(latitude_degree[4:], 2)}'
                    latitude_minutes = f'{human_watch_big(data_raw[2:3]):08b}'
                    coordinates += f'.{int(latitude_minutes[:4], 2)}{int(latitude_minutes[4:], 2)}'
                    buff_latitude_minutes_and_shares = f'{int(latitude_minutes[:4], 2)}{int(latitude_minutes[4:], 2)}'
                    latitude_shares = f'{human_watch_big(data_raw[3:4]):08b}'
                    coordinates += f"'{int(latitude_shares[:4], 2)}{int(latitude_shares[4:], 2)}"
                    buff_latitude_minutes_and_shares += f".{int(latitude_shares[:4], 2)}{int(latitude_shares[4:], 2)}"
                    latitude_shares_adddon_and_code_attitude = f'{human_watch_big(data_raw[4:5]):08b}'
                    buff_latitude_minutes_and_shares += f'{int(latitude_shares_adddon_and_code_attitude[:4], 2)}'
                    coordinates += f"{int(latitude_shares_adddon_and_code_attitude[:4], 2)}"
                    if (latitude_shares_adddon_and_code_attitude[-1] == '0'):
                        coordinates += f' с.ш. '
                        diraction_latitude = 'N'
                    else:
                        coordinates += f' ю.ш. '
                        diraction_latitude = 'S'
                    
                    longtitude_degree_senior_midl = f'{human_watch_big(data_raw[5:6]):08b}'
                    coordinates += f'{int(longtitude_degree_senior_midl[:4], 2)}{int(longtitude_degree_senior_midl[4:], 2)}'
                    longtitude_degree_junior_and_senior_minutes = f'{human_watch_big(data_raw[6:7]):08b}'
                    buff_longtitude_degree = f'{int(longtitude_degree_senior_midl[:4], 2)}{int(longtitude_degree_senior_midl[4:], 2)}{int(longtitude_degree_junior_and_senior_minutes[:4], 2)}'
                    coordinates += f'{int(longtitude_degree_junior_and_senior_minutes[:4], 2)}.{int(longtitude_degree_junior_and_senior_minutes[4:], 2)}'
                    longtitude_junior_minutes_and_dec_shares = f'{human_watch_big(data_raw[7:8]):08b}'
                    coordinates += f"{int(longtitude_junior_minutes_and_dec_shares[:4], 2)}'{int(longtitude_junior_minutes_and_dec_shares[4:], 2)}"
                    longtitude_shares_addon_and_code_longtitude = f'{human_watch_big(data_raw[8:9]):08b}'
                    coordinates += f"{int(longtitude_shares_addon_and_code_longtitude[:4], 2)}"
                    buff_longtitude_minutes_and_shares = f'{int(longtitude_degree_junior_and_senior_minutes[4:], 2)}{int(longtitude_junior_minutes_and_dec_shares[:4], 2)}.{int(longtitude_junior_minutes_and_dec_shares[4:], 2)}{int(longtitude_shares_addon_and_code_longtitude[:4], 2)}'
                    #print(buff_longtitude_degree, buff_longtitude_minutes_and_shares)
                    if (longtitude_shares_addon_and_code_longtitude[-1] == '0'):
                        coordinates += f' в.д. '
                        diraction_longtitude = 'E'
                    else:
                        coordinates += f' з.д. '
                        diraction_longtitude = 'W'
                
                    latitude_decode = dms_to_decimal(int(buff_latitude_degree), float(buff_latitude_minutes_and_shares), diraction_latitude)
                    longtitude_decode = dms_to_decimal(int(str(buff_longtitude_degree)), float(buff_longtitude_minutes_and_shares), diraction_longtitude)

                    coordinates += f' | {latitude_decode} {diraction_latitude}, {longtitude_decode} {diraction_longtitude}'

                    next_start_position = 9
                if (bits_presence[4] == '0'):
                    if (bits_presence[3] == '0'): # если счётчика отправленных пакетов нет и данных координат
                        count_input_packets = 'счётчика отправленных пакетов нет.'
                        next_start_position = 1 # пропускается (байтов нет)
                    else: # если счётчика отправленных пакетов нет, но данные координат есть
                        count_input_packets = 'счётчика отправленных пакетов нет.'
                        next_start_position = 9 
                else:
                    count_input_packets = human_watch(data_raw[next_start_position:next_start_position + 1]) # 1 байт на количество принятых пакетов
                    next_start_position = next_start_position + 1
                if (bits_presence[5] == '0'):
                    count_output_packets = 'счётчика принятых пакетов нет.'
                else:
                    count_output_packets = human_watch(data_raw[next_start_position:next_start_position + 1])
                    next_start_position = next_start_position + 1
                if (bits_presence[6] == '0'):
                    battery = 'информации о заряде батареи нет.'
                else:
                    battery_senior = human_watch(data_raw[next_start_position:next_start_position + 1])
                    battery_junior = human_watch(data_raw[next_start_position + 1: next_start_position + 2])
                    battery = f'{battery_senior}{battery_junior}'
                    battery = human_watch_big(data_raw[next_start_position:next_start_position + 2])
                    next_start_position = next_start_position + 2
                if (bits_presence[7] == '0'):
                    rssi = 'информации о RSSI нет.'
                    snr = 'информации о SNR нет.'
                else:
                    rssi = human_watch(data_raw[next_start_position:next_start_position + 1])
                    snr = human_watch(data_raw[next_start_position + 1: next_start_position + 2], True) # SNR в дополнительном коде (каком?)
                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=bits_presence)
                ws.cell(row=last_row+1, column=4, value=reason_send_packet)
                ws.cell(row=last_row+1, column=5, value=coordinates)
                ws.cell(row=last_row+1, column=6, value=count_input_packets)
                ws.cell(row=last_row+1, column=7, value=count_output_packets)
                ws.cell(row=last_row+1, column=8, value=battery)
                ws.cell(row=last_row+1, column=9, value=f'-{rssi}')
                ws.cell(row=last_row+1, column=10, value=snr)

            # Датчик движения Smart-MS0101 Вега
            elif (type_device == 'Smart-MS0101' and port == 2): # формат для данных, приходящих на 2 порт LoRaWAN!!!
                type_packet = human_watch(data_raw[:1])
                if (type_packet == 1):
                    type_packet_decode = 'текущее состояние'
                else:
                    type_packet_decode = f'неизвестный тип ({type_packet})'
                battery = human_watch(data_raw[1:2])
                main_settings = f'{human_watch(data_raw[2:3]):08b}'
                main_settings_decode = ''
                if (main_settings[0] == '0'):
                    main_settings_decode += 'Тип активации - OTAA\n'
                else:
                    main_settings_decode += 'Тип активации - ABP\n'
                if (main_settings[1] == '0'):
                    main_settings_decode += 'Запрос подтверждения пакетов - выключен\n'
                else:
                    main_settings_decode += 'Запрос подтверждения пакетов - включен\n'
                if (main_settings[2] == '0' and main_settings[3] == '0' and main_settings[4] == '0'):
                    main_settings_decode += 'Период выхода на связь - 5 минут'
                elif (main_settings[2] == '1' and main_settings[3] == '0' and main_settings[4] == '0'):
                    main_settings_decode += 'Период выхода на связь - 15 минут'
                elif (main_settings[2] == '0' and main_settings[3] == '1' and main_settings[4] == '0'):
                    main_settings_decode += 'Период выхода на связь - 30 минут'
                elif (main_settings[2] == '1' and main_settings[3] == '1' and main_settings[4] == '0'):
                    main_settings_decode += 'Период выхода на связь - 1 час'
                elif (main_settings[2] == '0' and main_settings[3] == '0' and main_settings[4] == '1'):
                    main_settings_decode += 'Период выхода на связь - 6 часов'
                elif (main_settings[2] == '1' and main_settings[3] == '0' and main_settings[4] == '1'):
                    main_settings_decode += 'Период выхода на связь - 12 часов'
                elif (main_settings[2] == '0' and main_settings[3] == '1' and main_settings[4] == '1'):
                    main_settings_decode += 'Период выхода на связь - 24 часа'
                # 5 6 7 бит - резерв
                temperature = human_watch(data_raw[3:5], True) / 10 # была умножена на 10
                reason_send_packet = human_watch(data_raw[5:6])
                if (reason_send_packet == 0):
                    reason_send_packet_decode = 'по времени'
                elif (reason_send_packet == 1):
                    reason_send_packet_decode = 'по тревоге'
                elif (reason_send_packet == 2):
                    reason_send_packet_decode = 'по автопостановке в охрану'
                unix_time = human_watch(data_raw[6:10])
                timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=type_packet_decode)
                ws.cell(row=last_row+1, column=4, value=battery)
                ws.cell(row=last_row+1, column=5, value=main_settings_decode)
                ws.cell(row=last_row+1, column=6, value=temperature)
                ws.cell(row=last_row+1, column=7, value=reason_send_packet_decode)
                ws.cell(row=last_row+1, column=8, value=timestamp)

            # Счётчик импульсов СИ-12 Вега
            elif (type_device == 'СИ-12' and port == 2): # формат для данных, приходящих на 2 порт LoRaWAN!!! (пока только для пакетов с типом 1 и 2!!!)
                type_packet = human_watch(data_raw[:1])
                if (type_packet == 1 or type_packet == 2):
                    if (type_packet == 1):
                        type_packet_decode = 'обычный пакет с показаниями'
                    else:
                        type_packet_decode = 'замыкание охранного входа'
                    battery = human_watch(data_raw[1:2])
                    main_settings = f'{human_watch(data_raw[2:3]):08b}'
                    main_settings_decode = ''
                    if (main_settings[0] == '0'):
                        main_settings_decode += 'Тип активации - OTAA\n'
                    else:
                        main_settings_decode += 'Тип активации - ABP\n'
                    if (main_settings[1] == '0' and main_settings[2] == '0' and main_settings[3] == '0'):
                        main_settings_decode += 'Период выхода на связь - 5 минут\n'
                    elif (main_settings[1] == '1' and main_settings[2] == '0' and main_settings[3] == '0'):
                        main_settings_decode += 'Период выхода на связь - 15 минут\n'
                    elif (main_settings[1] == '0' and main_settings[2] == '1' and main_settings[3] == '0'):
                        main_settings_decode += 'Период выхода на связь - 30 минут\n'
                    elif (main_settings[1] == '1' and main_settings[2] == '1' and main_settings[3] == '0'):
                        main_settings_decode += 'Период выхода на связь - 1 час\n'
                    elif (main_settings[1] == '0' and main_settings[2] == '0' and main_settings[3] == '1'):
                        main_settings_decode += 'Период выхода на связь - 6 часов\n'
                    elif (main_settings[1] == '1' and main_settings[2] == '0' and main_settings[3] == '1'):
                        main_settings_decode += 'Период выхода на связь - 12 часов\n'
                    elif (main_settings[1] == '0' and main_settings[2] == '1' and main_settings[3] == '1'):
                        main_settings_decode += 'Период выхода на связь - 24 часа\n'
                    if (main_settings[4] == '0'):
                        main_settings_decode += 'Тип первого входа - импульсный\n'
                    else:
                        main_settings_decode += 'Тип первого входа - охранный\n'
                    if (main_settings[5] == '0'):
                        main_settings_decode += 'Тип второго входа - импульсный\n'
                    else:
                        main_settings_decode += 'Тип второго входа - охранный\n'
                    if (main_settings[6] == '0'):
                        main_settings_decode += 'Тип третьего входа - импульсный\n'
                    else:
                        main_settings_decode += 'Тип третьего входа - охранный\n'
                    if (main_settings[7] == '0'):
                        main_settings_decode += 'Тип четвертого входа - импульсный'
                    else:
                        main_settings_decode += 'Тип четвертого входа - охранный'
                    unix_time = human_watch(data_raw[3:7])
                    timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                    temperature = human_watch(data_raw[7:8], True)

                    readings_entrance_one = human_watch(data_raw[8:12])
                    if (readings_entrance_one == 0 or readings_entrance_one == 1):
                        if (readings_entrance_one == 0):
                            if (main_settings[4] == '1'):
                                readings_entrance_one_decode = 'разомкнут'
                            else:
                                readings_entrance_one_decode = 'количество импульсов 0'
                        else:
                            if (main_settings[4] == '1'):
                                readings_entrance_one_decode = 'замкнут'
                            else:
                                readings_entrance_one_decode = 'количество импульсов 1'
                    else:
                        readings_entrance_one_decode = f'количество импульсов {readings_entrance_one}'
                    
                    readings_entrance_two = human_watch(data_raw[12:16])
                    if (readings_entrance_two == 0 or readings_entrance_two == 1):
                        if (readings_entrance_two == 0):
                            if (main_settings[5] == '1'):
                                readings_entrance_two_decode = 'разомкнут'
                            else:
                                readings_entrance_two_decode = 'количество импульсов 0'
                        else:
                            if (main_settings[5] == '1'):
                                readings_entrance_two_decode = 'замкнут'
                            else:
                                readings_entrance_two_decode = 'количество импульсов 1'
                    else:
                        readings_entrance_two_decode = f'количество импульсов {readings_entrance_two}'

                    readings_entrance_three = human_watch(data_raw[16:20])
                    if (readings_entrance_three == 0 or readings_entrance_three == 1):
                        if (readings_entrance_three == 0):
                            if (main_settings[6] == '1'):
                                readings_entrance_three_decode = 'разомкнут'
                            else:
                                readings_entrance_three_decode = 'количество импульсов 0'
                        else:
                            if (main_settings[6] == '1'):
                                readings_entrance_three_decode = 'замкнут'
                            else:
                                readings_entrance_three_decode = 'количество импульсов 1'
                    else:
                        readings_entrance_three_decode = f'количество импульсов {readings_entrance_three}'

                    readings_entrance_four = human_watch(data_raw[20:24])
                    if (readings_entrance_four == 0 or readings_entrance_four == 1):
                        if (readings_entrance_four == 0):
                            if (main_settings[7] == '1'):
                                readings_entrance_four_decode = 'разомкнут'
                            else:
                                readings_entrance_four_decode = 'количество импульсов 0'
                        else:
                            if (main_settings[7] == '1'):
                                readings_entrance_four_decode = 'замкнут'
                            else:
                                readings_entrance_four_decode = 'количество импульсов 1'
                    else:
                        readings_entrance_four_decode = f'количество импульсов {readings_entrance_four}'
                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=type_packet_decode)
                ws.cell(row=last_row+1, column=4, value=battery)
                ws.cell(row=last_row+1, column=5, value=main_settings_decode)
                ws.cell(row=last_row+1, column=6, value=timestamp)
                ws.cell(row=last_row+1, column=7, value=temperature)
                ws.cell(row=last_row+1, column=8, value=readings_entrance_one_decode)
                ws.cell(row=last_row+1, column=9, value=readings_entrance_two_decode)
                ws.cell(row=last_row+1, column=10, value=readings_entrance_three_decode)
                ws.cell(row=last_row+1, column=11, value=readings_entrance_four_decode)

            # Датчик климата/температуры/открытия Smart-HS0101 Вега
            elif (type_device == 'Smart-HS0101' and port == 2):
                type_packet = human_watch(data_raw[:1])
                if (type_packet == 1):
                    type_packet_decode = 'текущее состояние устройства'
                elif (type_packet == 2):
                    type_packet_decode = 'по датчику открытия 1'
                elif (type_packet == 3):
                    type_packet_decode = 'по датчику открытия 2'
                elif (type_packet == 4):
                    type_packet_decode = 'по акселерометру'
                elif (type_packet == 5):
                    type_packet_decode = 'по выходу влажности за установленные пороги'
                elif (type_packet == 6):
                    type_packet_decode = 'по выходу температуры за установленные пороги'
                else:
                    type_packet_decode = f'неизвестный тип ({type_packet})'
                battery = human_watch(data_raw[1:2])
                if (type_packet == 1):
                    unix_time = human_watch(data_raw[2:6])
                    timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                    timestamp_decode = f'снятие показаний в {timestamp}'
                else:
                    unix_time = human_watch(data_raw[2:6])
                    timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                    timestamp_decode = f'формирование пакета в {timestamp}'
                temperature = human_watch(data_raw[6:8], True) / 10 # в пакете умножена на 10
                humidity = human_watch(data_raw[8:9])
                state_sensor_one = human_watch(data_raw[9:10])
                if (state_sensor_one == 0):
                    state_sensor_one_decode = 'открыто'
                else:
                    state_sensor_one_decode = 'закрыто'
                state_sensor_two = human_watch(data_raw[10:11])
                if (state_sensor_two == 0):
                    state_sensor_two_decode = 'открыто'
                else:
                    state_sensor_two_decode = 'закрыто'
                ugol_otkl_vert = human_watch(data_raw[11:12])
                lower_threshold_humidity = human_watch(data_raw[12:13])
                upper_threshold_humidity = human_watch(data_raw[13:14])
                lower_threshold_temperature = human_watch(data_raw[14:15], True)
                upper_threshold_temperature = human_watch(data_raw[15:16], True)

                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=type_packet_decode)
                ws.cell(row=last_row+1, column=4, value=battery)
                ws.cell(row=last_row+1, column=5, value=timestamp_decode)
                ws.cell(row=last_row+1, column=6, value=temperature)
                ws.cell(row=last_row+1, column=7, value=humidity)
                ws.cell(row=last_row+1, column=8, value=state_sensor_one_decode)
                ws.cell(row=last_row+1, column=9, value=state_sensor_two_decode)
                ws.cell(row=last_row+1, column=10, value=ugol_otkl_vert)
                ws.cell(row=last_row+1, column=11, value=lower_threshold_humidity)
                ws.cell(row=last_row+1, column=12, value=upper_threshold_humidity)
                ws.cell(row=last_row+1, column=13, value=lower_threshold_temperature)
                ws.cell(row=last_row+1, column=14, value=upper_threshold_temperature)

            # Датчик ТД-11 Вега
            elif (type_device == 'ТД-11' and port == 2):
                type_packet = human_watch(data_raw[:1])
                if (type_packet == 0):
                    type_packet_decode = 'передача по времени'
                elif (type_packet == 1):
                    type_packet_decode = 'сработал охранный вход'
                elif (type_packet == 2):
                    type_packet_decode = 'сработал датчик вскрытия (тампер)'
                elif (type_packet == 3):
                    type_packet_decode = 'температура датчика NTC вышла за установленные пороги'
                elif (type_packet == 4):
                    type_packet_decode = 'превышен порог на импульсном входе'
                elif (type_packet == 5):
                    type_packet_decode = 'по запросу от сервера'
                else:
                    type_packet_decode = f'неизвестный тип ({type_packet})'
                battery = human_watch(data_raw[1:2])
                exceed_limits = human_watch(data_raw[2:3])
                if (exceed_limits == 0):
                    exceed_limits_decode = 'нет превышения'
                else:
                    exceed_limits_decode = 'есть превышение'
                unix_time = human_watch(data_raw[3:7])
                timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                temperature_device = human_watch(data_raw[7:8])
                temperature_ntc = human_watch(data_raw[8:10], True) / 10 # в пакете умножена на 10
                if (temperature_ntc == -100):
                    temperature_ntc = '[ВНИМАНИЕ] терморезистор не подключен к термодатчику (-100 *C)'
                elif (temperature_ntc == -127):
                    temperature_ntc = '[ВНИМАНИЕ] короткое замыкание терморезистора (-127 *C)'
                lower_threshold_temperature_sensor = human_watch(data_raw[10:11], True)
                upper_threshold_temperature_sensor = human_watch(data_raw[11:12], True)
                state_in = f'{human_watch(data_raw[16:17]):08b}'
                state_in_decode = ''
                if (state_in[0] == '0'):
                    state_in_decode += 'Состояние охранного входа: вход замкнут\n'
                else:
                    state_in_decode += 'Состояние охранного входа: вход разомкнут\n'
                if (state_in[1] == '0'):
                    state_in_decode += 'Состояние тампера (датчик вскрытия) : корпус не вскрыт\n'
                else:
                    state_in_decode += 'Состояние тампера (датчик вскрытия) : корпус вскрыт\n'
                if (state_in[2] == '0'):
                    state_in_decode += 'Состояние датчика Холла 1 (резерв): не задействован\n'
                else:
                    state_in_decode += f'Состояние датчика Холла 1 (резерв): неизвестное состояние ({state_in[2]})\n'
                if (state_in[3] == '0'):
                    state_in_decode += 'Состояние датчика Холла 2 (резерв): не задействован\n'
                else:
                    state_in_decode += f'Состояние датчика Холла 2 (резерв): неизвестное состояние ({state_in[3]})\n'
                if (state_in[5] == '0'):
                    state_in_decode += 'Режим входа: охранный (0)'
                else:
                    state_in_decode += 'Режим входа: импульсный (1)'
                indications_discr_in = human_watch(data_raw[12:16])
                if (state_in[5] == '0'):
                    indications_discr_in_decode = f'Показания на дискр. входе: {indications_discr_in} импульса(ов)'
                else:
                    if (indications_discr_in == 0):
                        indications_discr_in_decode = f'Показания на дискр. входе: вход разомкнут'
                    else:
                        indications_discr_in_decode = f'Показания на дискр. входе: вход замкнут'
                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=type_packet_decode)
                ws.cell(row=last_row+1, column=4, value=battery)
                ws.cell(row=last_row+1, column=5, value=exceed_limits_decode)
                ws.cell(row=last_row+1, column=6, value=timestamp)
                ws.cell(row=last_row+1, column=7, value=temperature_device)
                ws.cell(row=last_row+1, column=8, value=temperature_ntc)
                ws.cell(row=last_row+1, column=9, value=lower_threshold_temperature_sensor)
                ws.cell(row=last_row+1, column=10, value=upper_threshold_temperature_sensor)
                ws.cell(row=last_row+1, column=11, value=state_in_decode)
                ws.cell(row=last_row+1, column=12, value=indications_discr_in_decode)

            
            elif (type_device == 'Smart-MC0101' and port == 2):
                type_packet = human_watch(data_raw[:1])
                battery = human_watch(data_raw[1:2])
                temperature = human_watch(data_raw[3:5], True) / 10
                reason_send_packet = human_watch(data_raw[5:6])
                if (reason_send_packet == 0):
                    reason_send_packet_decode = 'по времени'
                elif (reason_send_packet == 1):
                    reason_send_packet_decode = 'сработал датчик открытия 1'
                elif (reason_send_packet == 2):
                    reason_send_packet_decode = 'сработал датчик открытия 2'
                else:
                    reason_send_packet_decode = 'неизвестная причина'
                state_in = f'{human_watch(data_raw[6:7]):08b}'
                state_in_decode = ''
                if (state_in[0] == 1):
                    state_in_decode += 'Датчик открытия 1 (состояние магнита) - поднесен\n'
                else:
                    state_in_decode += 'Датчик открытия 1 (состояние магнита) - не поднесен\n'
                if (state_in[1] == 1):
                    state_in_decode += 'Датчик открытия 2 (состояние магнита) - поднесен'
                else:
                    state_in_decode += 'Датчик открытия 2 (состояние магнита) - не поднесен'
                unix_time = human_watch(data_raw[7:11])
                timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=type_packet)
                ws.cell(row=last_row+1, column=4, value=battery)
                ws.cell(row=last_row+1, column=5, value=temperature)
                ws.cell(row=last_row+1, column=6, value=reason_send_packet_decode)
                ws.cell(row=last_row+1, column=7, value=state_in_decode)
                ws.cell(row=last_row+1, column=8, value=timestamp)

            elif (type_device == 'Smart-SS0102' and port == 2):
                type_packet = human_watch(data_raw[:1])
                if (type_packet == 5):
                    type_packet_decode = 'тревога по обнаружению пожара'
                elif (type_packet == 6):
                    type_packet_decode = 'тест'
                elif (type_packet == 7):
                    type_packet_decode = 'тревога по солидарной линии работы'
                elif (type_packet == 8):
                    type_packet_decode = 'снятие с крепежной платформы'
                elif (type_packet == 9):
                    type_packet_decode = 'сброс тревоги'
                elif (type_packet == 10):
                    type_packet_decode = 'низкий заряд АКБ'
                elif (type_packet == 11):
                    type_packet_decode = 'старт охраны датчика'
                elif (type_packet == 12):
                    type_packet_decode = 'стоп охраны датчика'
                elif (type_packet == 13):
                    type_packet_decode = 'резерв'
                elif (type_packet == 14):
                    type_packet_decode = 'ошибка датчика'
                elif (type_packet == 15):
                    type_packet_decode = 'запыленность камеры датчика'
                elif (type_packet == 6):
                    type_packet_decode = 'данные по расписанию'
                unix_time = human_watch(data_raw[1:5])
                timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                current_status = f'{human_watch(data_raw[5:6]):08b}'
                current_status_decode = ''
                if (current_status[0] == 1):
                    current_status_decode += 'запыленность\n'
                if (current_status[1] == 1):
                    current_status_decode += 'пожар\n'
                if (current_status[2] == 1):
                    current_status_decode += 'тест\n'
                if (current_status[3] == 1):
                    current_status_decode += 'дежурный подрежим\n'
                if (current_status[4] == 1):
                    current_status_decode += 'неисправность\n'
                if (current_status[5] == 1):
                    current_status_decode += 'тревога\n'
                if (current_status[6] == 1):
                    current_status_decode += 'снятие с крепежной платформы\n'
                if (current_status[7] == 1):
                    current_status_decode += 'сигнал по линии солидарной работы\n'
                voltage = human_watch(data_raw[6:8])
                amperage = human_watch(data_raw[8:10])
                temperature_termistor = human_watch(data_raw[10:12])
                flag_use_battery_one = human_watch(data_raw[12:13])
                if (flag_use_battery_one == 1):
                    flag_use_battery_one_decode = 'используется'
                else:
                    flag_use_battery_one_decode = 'не используется'
                flag_use_battery_two = human_watch(data_raw[13:14])
                if (flag_use_battery_two == 1):
                    flag_use_battery_two_decode = 'используется'
                else:
                    flag_use_battery_two_decode = 'не используется'
                flag_presence_battery_one = human_watch(data_raw[14:15])
                if (flag_presence_battery_one == 1):
                    flag_presence_battery_one_decode = 'присутствует'
                else:
                    flag_presence_battery_one_decode = 'отсутствует'
                flag_presence_battery_two = human_watch(data_raw[15:16])
                if (flag_presence_battery_two == 1):
                    flag_presence_battery_two_decode = 'присутствует'
                else:
                    flag_presence_battery_two_decode = 'отсутствует'
                battery_one = human_watch(data_raw[16:17])
                battery_two = human_watch(data_raw[17:18])

                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=type_packet_decode)
                ws.cell(row=last_row+1, column=4, value=timestamp)
                ws.cell(row=last_row+1, column=5, value=current_status_decode)
                ws.cell(row=last_row+1, column=6, value=voltage)
                ws.cell(row=last_row+1, column=7, value=amperage)
                ws.cell(row=last_row+1, column=8, value=temperature_termistor)
                ws.cell(row=last_row+1, column=9, value=flag_use_battery_one)
                ws.cell(row=last_row+1, column=10, value=flag_use_battery_two)
                ws.cell(row=last_row+1, column=11, value=flag_presence_battery_one)
                ws.cell(row=last_row+1, column=12, value=flag_presence_battery_two)
                ws.cell(row=last_row+1, column=13, value=battery_one)
                ws.cell(row=last_row+1, column=14, value=battery_two)
            elif (type_device == 'Smart-UM0101' and port == 2):
                type_packet = human_watch(data_raw[0:1])
                if (type_packet == 1):
                    type_packet_decode = 'текущее состояние устройства'
                elif (type_packet == 2):
                    type_packet_decode = 'по выходу CO2 за установленные пороги'
                elif (type_packet == 3):
                    type_packet_decode = 'по выходу уровня освещенности за установленые пороги'
                elif (type_packet == 4):
                    type_packet_decode = 'по акселерометру (резерв)'
                elif (type_packet == 5):
                    type_packet_decode = 'по выходу влажности за установленные пороги'
                elif (type_packet == 6):
                    type_packet_decode = 'по выходу температуры за установленные пороги'
                elif (type_packet == 7):
                    type_packet_decode = 'по выходу уровня шума за определенные пороги'
                elif (type_packet == 8):
                    type_packet_decode = 'при обнаружении снятия'
                else:
                    type_packet_decode = 'неизвестный тип'
                battery = human_watch(data_raw[1:2])
                unix_time = human_watch(data_raw[2:6])
                timestamp = datetime.fromtimestamp(unix_time * 1e-3)
                state_power = human_watch(data_raw[6:7])
                if (state_power == 1):
                    state_power_decode = 'от батарей'
                else:
                    state_power_decode = 'внешнее'
                temperature = human_watch(data_raw[7:9], True) / 10
                humidity = human_watch(data_raw[9:10])
                light = human_watch(data_raw[10:12])
                noise = human_watch(data_raw[12:13])
                co2 = human_watch(data_raw[13:15])
                ugol_otkl_vert = human_watch(data_raw[15:16])
                lower_threshold_temperature = human_watch(data_raw[16:17], True)
                upper_threshold_temperature = human_watch(data_raw[17:18], True)
                lower_threshold_humidity = human_watch(data_raw[18:19])
                upper_threshold_humidity = human_watch(data_raw[19:20])
                lower_threshold_light = human_watch(data_raw[20:21])
                upper_threshold_light = human_watch(data_raw[21:22])
                lower_threshold_noise = human_watch(data_raw[22:23])
                upper_threshold_noise = human_watch(data_raw[23:24])
                lower_threshold_co2 = human_watch(data_raw[24:25])
                upper_threshold_co2 = human_watch(data_raw[25:26])

                ws = wb[f'{deveui} ({type_device})']
                ws.cell(row=1, column=2, value=f'{deveui}')
                ws.cell(row=1, column=4, value=f'{type_device}')
                ws.cell(row=1, column=5, value=f'{classroom}')
                last_row = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None and cell != '' for cell in row):
                        last_row += 1
                ws.cell(row=last_row+1, column=1, value=timestamp_bd)
                ws.cell(row=last_row+1, column=2, value=macbs)
                ws.cell(row=last_row+1, column=3, value=type_packet_decode)
                ws.cell(row=last_row+1, column=4, value=battery)
                ws.cell(row=last_row+1, column=5, value=timestamp)
                ws.cell(row=last_row+1, column=6, value=state_power_decode)
                ws.cell(row=last_row+1, column=7, value=temperature)
                ws.cell(row=last_row+1, column=8, value=humidity)
                ws.cell(row=last_row+1, column=9, value=light)
                ws.cell(row=last_row+1, column=10, value=noise)
                ws.cell(row=last_row+1, column=11, value=co2)
                ws.cell(row=last_row+1, column=12, value=ugol_otkl_vert)
                ws.cell(row=last_row+1, column=13, value=lower_threshold_temperature)
                ws.cell(row=last_row+1, column=14, value=upper_threshold_temperature)
                ws.cell(row=last_row+1, column=15, value=lower_threshold_humidity)
                ws.cell(row=last_row+1, column=16, value=upper_threshold_humidity)
                ws.cell(row=last_row+1, column=17, value=lower_threshold_light)
                ws.cell(row=last_row+1, column=18, value=upper_threshold_light)
                ws.cell(row=last_row+1, column=19, value=lower_threshold_noise)
                ws.cell(row=last_row+1, column=20, value=upper_threshold_noise)
                ws.cell(row=last_row+1, column=21, value=lower_threshold_co2)
                ws.cell(row=last_row+1, column=22, value=upper_threshold_co2)



            # не поддерживаемые устройства
            elif (type_device not in support_device_list):
                if (settings_app[0] == 1):
                    print(f'[Ошибка в отчёте {timestamp_bd}]: К сожалению, данное устройство: "{deveui}" ({type_device}) пока не поддерживается\n= = =')

            else:
                if (settings_app[0] == 1):
                    print(f'[Ошибка в отчёте {timestamp_bd}]: К сожалению, данный вид отчёта от устройства типа: "{type_device}" приходящий на порт: "{port}" пока не поддерживается\n= = =')





            # демонстрация обработанных данных в зависимости от deveui и частных случаев

            if settings_app[3] != 0: # для отображения только нового трафика
                if (not(timestamp_bd > start_time)):
                    continue

            if settings_app[2] != '' and deveui != settings_app[2]: # для фильтрации только нужного трафика
                continue

            is_new = timestamp_bd > start_time
            note = f"[НОВЫЙ] " if is_new else ""

            # демонстрация ТС-12 Вега
            if (type_device == 'ТС-12' and port == 4):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nБиты присутствия: {bits_presence}\nПричина отправки пакета: {reason_send_packet}\nНавигационные данные: {coordinates}\nКоличество принятых пакетов: {count_input_packets}\nКоличество отправленных пакетов: {count_output_packets}\nЗаряд батареи (мВ): {battery}\nRSSI: -{rssi} dBm\nSNR: {snr} dB\n= = =')
            
            # демонстрация Smart-MS0101 Вега
            elif (type_device == 'Smart-MS0101' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet_decode}\nЗаряд батареи: {battery}%\n[Основные параметры (начало)]:\n{main_settings_decode}\n[Основные параметры (конец)]\nТемпература: {temperature} *C\nПричина отправки пакета: {reason_send_packet_decode}\nВремя: {timestamp}\n= = =')
            
            # демонстрация СИ-12 Вега
            elif (type_device == 'СИ-12' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet_decode}\nЗаряд батареи: {battery}%\n[Основные параметры (начало)]:\n{main_settings_decode}\n[Основные параметры (конец)]\nВремя: {timestamp}\nТемпература: {temperature} *C\n[Показания входов]\n[[Вход 1]]\n{readings_entrance_one_decode}\n[[Вход 2]]\n{readings_entrance_two_decode}\n[[Вход 3]]\n{readings_entrance_three_decode}\n[[Вход 4]]\n{readings_entrance_four_decode}\n= = =')
            
            # демонстрация Smart-HS0101 Вега
            elif (type_device == 'Smart-HS0101' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet_decode}\nЗаряд батареи: {battery}%\nВремя: {timestamp_decode}\nТемпература: {temperature} *C\nВлажность: {humidity}%\n[Состояние датчиков открытия]\n[[Датчик 1]]\n{state_sensor_one_decode}\n[[Датчик 2]]\n{state_sensor_two_decode}\nУгол отклонения от вертикали: {ugol_otkl_vert}*\n[Пороговые значения влажности]: {lower_threshold_humidity}% - {upper_threshold_humidity}%\n[Пороговые значения температуры]: {lower_threshold_temperature} *C - {upper_threshold_temperature} *C\n= = =')
            
            # демонстрация Smart Badge Вега
            elif (type_device == 'Smart Badge' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet_decode}\nПричина формирования пакета: {reason_format_packet_decode}\nЗаряд батареи: {battery}%\nВремя: {timestamp}\nТемпература: {temperature} *C\n[Состояние бейджа (начало)]:\n{state_string}\n[Состояние бейджа (конец)]\nУгол отклонения от вертикали: {ugol_otkl_vert}*')
                if (type_packet == 1):
                    print(f'Координаты (широта, долгота, высота над ср. уровнем моря): {latitude}, {longtitude}, {height} м.\nКурс: {course}*\nСкорость: {speed} км/ч\nКоличество видимых спутников: {count_sputnik}\nКоличество спутников решения: {count_decision_sputnik}\nСостояние меток СИЗ: {state_tag_SIZ_decode}\n= = =')
                elif (type_packet == 2):
                    print(f'Тип BLE-маяка: {type_ble_beacon}\nНаименование BLE-маяка и его параметры: {ble_beacon}\nЭталонное значение RSSI: {rssi_etalon}\nЗначение TX_POWER: {tx_power}\nСостояние меток СИЗ: {state_tag_SIZ_decode}\n= = =')
                elif (type_packet == 5):
                    print(f'[BLE-метка №1]\nMAC: {mac_tag_one}\nЗаряд батареи: {battery_tag_one}%\nТемпература: {temperature_tag_one} *C\nВлажность: {humidity_tag_one}%\nЭталонный RSSI: {rssi_etalon_tag_one}\nЗначение TX_POWER: {tx_power_tag_one}\n\n[BLE-метка №2]\nMAC: {mac_tag_two}\nЗаряд батареи: {battery_tag_two}%\nТемпература: {temperature_tag_two} *C\nВлажность: {humidity_tag_two}%\nЭталонный RSSI: {rssi_etalon_tag_two}\nЗначение TX_POWER: {tx_power_tag_two}\n\n[BLE-метка №3]\nMAC: {mac_tag_three}\nЗаряд батареи: {battery_tag_three}%\nТемпература: {temperature_tag_three} *C\nВлажность: {humidity_tag_three}%\nЭталонный RSSI: {rssi_etalon_tag_three}\nЗначение TX_POWER: {tx_power_tag_three}\n\nСостояние меток СИЗ: {state_tag_SIZ_decode}\n= = =')
            
            # демонстрация Smart-WB0101 Вега
            elif (type_device == 'Smart-WB0101' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nРежим работы: {operating_mode_decode}\nЗаряд батареи: {battery}%\nВремя: {timestamp}\nТемпература: {temperature} *C\n= = =')
            
            # демонстрация ТД-11 rev2 Вега
            elif (type_device == 'ТД-11' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet_decode}\nЗаряд батареи: {battery}%\nПроверка лимитов: {exceed_limits_decode}\nВремя: {timestamp}\nТемпература устройства: {temperature_device} *C\nТемпература датчика NTC: {temperature_ntc} *C\nНижний предел температуры датчика: {lower_threshold_temperature_sensor} *C\nВерхний предел температуры датчика: {upper_threshold_temperature_sensor} *C\n[Состояние входа (начало)]\n{state_in_decode}\n[Состояние входа (конец)]\nПоказатели на дискретном входе: {indications_discr_in_decode}\n= = =')

            # демонстрация Smart-MC0101 Вега
            elif (type_device == 'Smart-MC0101' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet}\nЗаряд батареи: {battery}%\nТемпература: {temperature} *C\nПричина отправки: {reason_send_packet_decode}\nСостояние входов:\n{state_in_decode}\nВремя снятия показаний/Время формирования пакета: {timestamp}\n= = =')
        
            # демонстрация Smart-SS0102 Вега
            elif (type_device == 'Smart-SS0102' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet_decode}\nВремя формирования пакета: {timestamp}\nТекущий статус: {current_status_decode}\nНапряжение с приемника: {voltage} мВ\nТок передатчика: {amperage} мА\nТемпература на термисторе: {temperature_termistor} *C\nФлаг - используется батарея 1: {flag_use_battery_one_decode}\nФлаг - используется батарея 2: {flag_use_battery_two_decode}\nФлаг присутствия батареи 1: {flag_presence_battery_one_decode}\nФлаг присутствия батареи 2: {flag_presence_battery_two_decode}\nЗаряд батареи 1: {battery_one}%\nЗаряд батареи 2: {battery_two}%\n= = =')
        
            # демонстрация Smart-UM0101
            elif (type_device == 'Smart-UM0101' and port == 2):
                print(f'= = =\n{note}[Отчёт {timestamp_bd}]:\n[ОСНОВНОЕ]:\nТип устройства: {type_device}\nКабинет: {classroom}\nПорт: {port}\nDEVEUI: {deveui}\nПоступило с БС: {macbs}')
                print(f'[ПЕРЕДАНО]:\nТип пакета: {type_packet_decode}\nЗаряд батареи: {battery}%\nВремя снятия показаний: {timestamp}\nСостояние питания: {state_power_decode}\nТемпература: {temperature} *C\nВлажность: {humidity} %\nУровень освещенности: {light}\nУровень шума: {noise}\nУровень CO2: {co2} ppm\nУгол отклонения от вертикали: {ugol_otkl_vert}\nНижний порог температуры: {lower_threshold_temperature} *C\nВерхний порог температуры: {upper_threshold_temperature} *C\nНижний порог влажности: {lower_threshold_humidity} %\nВерхний порог влажности: {upper_threshold_humidity} %\nНижний порог уровня освещенности: {lower_threshold_light}\nВерхний порог уровня освещенности: {upper_threshold_light}\nНижний порог уровня шума: {lower_threshold_noise}\nВерхний порог уровня шума: {upper_threshold_noise}\nНижний порог уровня CO2: {lower_threshold_co2} ppm\nВерхний порог уровня CO2: {upper_threshold_co2} ppm\n= = =')

        # эксклюзивное оповещение для первого чтения БД
        if (count == 0):
            if (settings_app[3] == 0):
                print('[Процесс обработки]: Обработка уже полученных ранее отчётов завершена.')
            print('\n[Процесс обработки]: Включено прослушивание новых поступающих отчётов.')
        count += 1

except KeyboardInterrupt:
    print('\n[Процесс обработки]: Выключено прослушивание новых поступающих отчётов, процесс сохранения файла отчётов и завершение программы.')
    while True:
        try:
            # Получаем путь к текущей директории, где находится .exe
            if getattr(sys, 'frozen', False):  # Если приложение запущено как .exe
                script_dir = Path(sys.executable).parent
            else:  # Если запущено как скрипт
                script_dir = Path(__file__).parent
            current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            # Указываем имя файла для сохранения
            output_file = script_dir / f"отчёт_за_{current_time}.xlsx"
            wb.save(output_file)
            print(f"Файл сохранён: {output_file.resolve()}")
            break
        except PermissionError:
            input("Ошибка: невозможно сохранить файл. Возможно, он уже открыт или нет прав на запись.\nПопробовать повторно: ")
    sleep(1)
for i in range(3):
    if (i == 0):
        print(f'Программа закроется через 3 секунды', end="\r")
    if (i == 1):
        print(f'Программа закроется через 2 секунды',end="\r")
    if (i == 2):
        print(f'Программа закроется через 1 секунду', end="\r")
    sleep(1)
# закрытие проводника по базе данных
cursor.close()
# закрытие подключения
connection.close()